from django import forms
from django.forms import ModelForm, TextInput, Textarea, Select, IntegerField, modelformset_factory
from django.forms import BaseModelFormSet, BooleanField, ValidationError
from django.shortcuts import reverse
from django.core.files.storage import FileSystemStorage
from django.utils import timezone
from crispy_forms.helper import FormHelper
from crispy_forms.layout import Layout, Submit, Row, ButtonHolder, Button
from .models import User, Quiz, QuizRoom, Question, QuizRoomMember, Answer
from .models import QuizRoomMember, RoomMemberStatus, CheckStatus, TransitionType, RoomStatus
from django.db.models import Max
from .constants import RoomStatus as rs, ExcelConstants as xl, RoomMemberStatus as rms, CheckStatus as cs
from .constants import FormFunction as ff, AnimationType as at, DefaultImages as di, MessageTypes as mt
from .functions import add_custom_field_error, set_time_limit, set_score_multiplier, add_custom_non_field_alert
from .functions import get_new_question_number, get_icon_buttons, get_last_question_number
from .functions import get_url_content_type, generate_room, isEmpty, upload_media_files, get_upload_media_filenames
from synaptic.classes.synaptic.CustomComponents import MultipleFileInput
from django.utils.translation import gettext_lazy as _
from datetime import datetime as dt
import openpyxl
import os
import json
import urllib

DisplayMoneyAttrs = {'class': 'integer', "maxlength": "8"}
DisplayPercentAttrs = {'disabled': True, 'class': 'bg-white percent'}

class CreateAnswerForm(ModelForm):
    class Meta:
        model = Answer;
        fields = ['answer', 'correct_answer']
        widgets = {
            'answer': TextInput(attrs={"placeholder": "Answer text"})
        }
    def __init__(self, *args, **kwargs):
        self.request = kwargs.pop('request', None)
        super().__init__(*args, **kwargs)
        for field in ['answer']:
            self.fields[field].label = ""

class JoinRoomForm(ModelForm):
    nickname = forms.CharField(
        label="Nickname", required=False, max_length=30)

    class Meta:
        model = QuizRoom;
        fields = ['room_number']
        widgets = {'room_number': TextInput(attrs=DisplayMoneyAttrs)}

    def __init__(self, *args, **kwargs):
        self.request = kwargs.pop('request', None)
        super().__init__(*args, **kwargs)
        self.helper = FormHelper(self)
        self.helper.layout = Layout(
            Row('room_number',
                'nickname',
                Submit('submit', 'Join', css_class='col-md-1 btn btn-primary'),
                css_class="row col-2 text-center justify-content-center mx-auto")
        )

    def clean(self):
        self.cleaned_data = super().clean()
        if not len(self.errors) == 0:
            return
        nickname = self.cleaned_data['nickname']
        if nickname == '':
            self.add_error('nickname', _('Nickname must be entered'))

        user = User.objects.get(username=self.request.user.username)
        entered_room_number = self.cleaned_data['room_number']
        room = QuizRoom.objects.get_or_none(room_number=entered_room_number)
        if room is None:
            self.add_error('room_number', _('Room not found'))
            return
        if room.status.description != rs.WAITING:
            room_member = QuizRoomMember.objects.get_or_none(room=room, user=user)
            if room_member is None:
                self.add_error('room_number', _('Room is not open for joining'))
                return

        room_members = QuizRoomMember.objects.filter(room=room, nickname=nickname).exclude(user=user)
        if len(room_members) > 0:
            self.add_error('nickname', _('Nickname is already being used'))


    def get_room_member(self, request):
        user = User.objects.get(username=request.user)
        member_rooms = QuizRoomMember.objects.filter(user=user).exclude(room__host=user).order_by("-joined_date")
        return member_rooms

    def save_room_member(self, form, user):
        room_number = form.cleaned_data["room_number"]
        room = QuizRoom.objects.get_or_none(room_number=room_number)
        status = RoomMemberStatus.objects.get_or_none(description=rms.LEFT)
        room_member = QuizRoomMember.objects.get_or_none(room=room, user=user)
        if room_member is None:
            room_member = QuizRoomMember(room=room, user=user, status=status, preview=False)
        nickname = form.cleaned_data['nickname']
        room_member.nickname = nickname
        room_member.joined_date = timezone.now()
        room_member.save()

class ListQuestionForm(forms.Form):
    question_id = IntegerField(required=True)
    changed = IntegerField(required=True)
    deleted = IntegerField(required=True)
    new_question_number = IntegerField(required=True)

    def __init__(self, *args, **kwargs):
        self.request = kwargs.pop('request', None)
        self.quiz = kwargs.pop('quiz', None)
        self.changed = False
        self.deleted = False
        super().__init__(*args, **kwargs)
    def clean(self):
        self.cleaned_data = super().clean()
        if not len(self.errors) == 0:
            return
        question_id = self.cleaned_data.get('question_id')
        if self.cleaned_data.get('changed'):
            self.changed = True
        if self.cleaned_data.get('deleted'):
            self.deleted = True
        if not self.changed and not self.deleted:
            self.question = None
            return
        self.question = Question.objects.get_or_none(pk=question_id, quiz=self.quiz)
        self.question.question_number = self.cleaned_data.get('new_question_number')
        if self.question is None:
            self.add_error('question_id', _('Question record does not match question data'))
            return

    @staticmethod
    def get_update_preview_redirect(request, quiz):
        resp = {}
        selected_question = request.POST.get('selected_question_number')
        if selected_question is not None:
            request.session["question_number"] = selected_question
            start_question = Question.objects.get_or_none(quiz=quiz, question_number = selected_question)
            if start_question == None:
                add_custom_non_field_alert(resp, "Unable to update selected question")
                return resp, None
            last_question = Question.objects.filter(quiz=quiz).last()
            if last_question == None:
                last_question = start_question
            request.session['return_to_url'] = "list_questions"
            button = request.POST.get('submit')
            if button == "Update":
                request.session["question_action"] = ff.UPDATE
                resp['redirect'] = reverse("synaptic:question")
            if button == "Preview":
                room = generate_room(quiz.pk, status=rs.QUESTION_PREVIEW, preview=True,
                                     first_question=start_question, last_question=last_question)
                room_member_status = RoomMemberStatus.objects.get_or_none(description=rms.JOINED)
                QuizRoomMember(room=room, user=quiz.created_by, status=room_member_status, preview=True).save()
                resp['redirect'] = reverse("synaptic:live_room", args=[room.room_number])
        return resp

    @staticmethod
    def process_add_button(request, resp):
        request.session['question_action'] = ff.CREATE
        request.session["question_number"] = None
        resp['redirect'] = reverse("synaptic:question")
        resp['success'] = True
        return resp

    @staticmethod
    def process_back_button(resp):
        resp['redirect'] = reverse("synaptic:list_quizzes")
        resp['success'] = True
        return resp

    @staticmethod
    def process_cancel_button(resp):
        resp['redirect'] = reverse("synaptic:list_questions")
        resp['success'] = True
        return resp

    @staticmethod
    def process_preview_button(request, resp):
        quiz_pk = request.session.get('quiz_pk')
        first_question_number = request.POST['selected_question_number']
        room = generate_room(
            quiz_pk, status=rs.QUESTION_PREVIEW, preview=True,
            first_question_number=first_question_number, last_question_number=first_question_number)
        request.session['preview_return_to_url'] = "list_questions"
        resp['redirect'] = reverse("synaptic:live_room", args=[room.room_number])

    @staticmethod
    def process_update_button(request, resp):
        request.session["question_action"] = ff.UPDATE
        request.session["question_number"] = int(request.POST['selected_question_number'])
        resp['redirect'] = reverse("synaptic:question")
        resp['success'] = True

    @staticmethod
    def save_updates(quiz, questions_data, request, resp):
        changed = False
        list_question_form = None
        update_objects, delete_ids = [], []
        for data in questions_data:
            list_question_form = ListQuestionForm(data=data, request=request, quiz=quiz)
            if list_question_form.is_valid():
                if list_question_form.question is not None:
                    if list_question_form.changed:
                        update_objects.append(list_question_form.question)
                        changed = True
                    if list_question_form.deleted:
                        delete_ids.append(list_question_form.question.pk)
                        changed = True

            else:
                add_custom_non_field_alert(resp, "Unable to save")
                return

        Question.objects.filter(pk__in=delete_ids).delete()
        Question.objects.bulk_update(update_objects, ['question_number'])
        if changed:
            request.session['message'] = {"message": "Updates Saved", "message_type": mt.SUCCESS}

        return list_question_form

    @staticmethod
    def validate_ajax_questions_data(request, resp):
        questions_data_raw = request.POST.get('questionsData')
        if questions_data_raw is None:
            add_custom_non_field_alert(resp, "Unable to save - no question data")
            return None

        button = request.POST['submit']

        if button in ['Preview', 'Update']:
            selected_question = request.POST.get('selected_question_number')
            if selected_question is None:
                add_custom_non_field_alert(resp, "Unable to save - question data invalid")
                return None

        questions_data = json.loads(questions_data_raw)
        if type(questions_data) != list:
            add_custom_non_field_alert(resp, "Unable to save - question data invalid")
            return None
        new_question_numbers = []
        for data in questions_data:
            if type(data) == dict:
                new_question_number = data.get('new_question_number')
                if new_question_number is None:
                    add_custom_non_field_alert(resp, "Unable to save - missing question number")
                    return None
                if new_question_number != int(new_question_number) or new_question_number < 1:
                    add_custom_non_field_alert(resp, "Unable to save - invalid new_question_number")
                    return None
                if not data.get('deleted'):
                    new_question_numbers.append(new_question_number)
            else:
                add_custom_non_field_alert(resp, "Unable to save - invalid question data content")
                return questions_data

        if len(new_question_numbers) != len(set(new_question_numbers)):
            add_custom_non_field_alert(resp, "Unable to save - duplicate question numbers")
            return questions_data

        for i, new_question_number in enumerate(sorted(new_question_numbers)):
            if new_question_number != i + 1:
                add_custom_non_field_alert(resp, "Unable to save - question numbers not in sequence")
                return questions_data

        del resp['success']
        return questions_data

    @staticmethod
    def validate_and_save(request, resp):
        questions_data = ListQuestionForm.validate_ajax_questions_data(request, resp)
        if 'errors' in resp:
            return None, None

        user, quiz = ListQuestionForm.validate_session_vars(request, resp)
        if 'errors' in resp:
            return None, None

        list_question_form = ListQuestionForm.save_updates(quiz, questions_data, request, resp)
        if 'errors' in resp:
            return None, None

        return list_question_form, quiz

    @staticmethod
    def validate_button(request, resp):
        if request.POST.get('submit') not in ['Add', 'Back', 'Cancel', 'Save', 'Update', 'Preview']:
            add_custom_non_field_alert(resp, "Invalid submission received")
        return resp

    @staticmethod
    def validate_session_vars(request, resp):
        user = User.objects.get(username=request.user.username)
        quiz_id = request.session.get('quiz_pk')
        quiz = Quiz.objects.get_or_none(pk=quiz_id, created_by=user)
        if quiz is None:
            resp['success'] = False
            add_custom_non_field_alert(resp, "Unable to save")
            return None, None
        return user, quiz

class ListQuizForm(forms.ModelForm):
    deleted = BooleanField(required=False,
                           widget=forms.CheckboxInput(attrs={'hidden': True})
                           )
    class Meta:
        model = Quiz
        fields=['id', 'title', 'created_date', 'status']
        widgets={
            'id': TextInput(attrs={'class': 'position: absolute', 'hidden': True}),
            'title': TextInput(attrs={'class': 'font-colour-pri'}),
            'status': TextInput(attrs={'readonly': 'readonly'}),
            'created_date': TextInput(attrs={'readonly': 'readonly'})
            }
    def __init__(self, *args, **kwargs):
        self.request = kwargs.pop('request', None)
        super().__init__(*args, **kwargs)
        self.field_custom_errors = {}
        self.non_field_custom_errors = {}
        self.non_field_custom_errors['resp'] = {}
    # def add_custom_field_error(self, field, error_msg, target_css_id):
    #     # used to add errors when the errors are going to be passed to and rendered by js
    #     # field must be a valid field for the form
    #     # error is the error message to be displayed
    #     # target_field is the css id of the field to which the error message will be appended
    #     # self.field_custom_errors[target_css_id] = add_custom_field_error(self.field_custom_errors, error, target_css_id)
    #     self.field_custom_errors[target_css_id] = error_msg
    #     self.add_error(field, error_msg)
    def add_custom_non_field_error(self, error_msg):
        add_custom_non_field_alert(self.non_field_custom_errors['resp'], error_msg)
        raise forms.ValidationError(_(error_msg))
    def clean(self):
        self.cleaned_data = super().clean()
        if not len(self.errors) == 0:
            return
        id = self.cleaned_data.get("id")
        if id is None or not hasattr(id, "id"):
            self.add_custom_non_field_error("Id error")
            #self.add_custom_field_error('id', 'Id error', 'success_message')
            # self.add_error('id', _('Id error'))
            return
        if self.cleaned_data['title'] == "":
            self.add_custom_non_field_error("Title must be entered")
            #self.add_custom_field_error('title', 'Title must be entered', 'success_message')
            # self.add_error('id', _('Id error'))
            return

class BaseListQuizFormSet(BaseModelFormSet):
    def __init__(self, *args, **kwargs):
        self.request = kwargs.pop('request', None)
        super().__init__(*args, **kwargs)
        self.field_custom_errors = {}

    def raise_non_field_error(self, error_msg):
        # used to add errors when the errors are going to be passed to and rendered by js in the message bar
        add_custom_non_field_alert(self.request.session['resp'], error_msg)
        raise forms.ValidationError(_(error_msg))

    def clean(self):
        # self.cleaned_data = super().clean()
        super(BaseListQuizFormSet, self).clean()
        if self.total_error_count() > 0:
            return
        received_ids, db_ids = [], []
        self.quizzes, self.received_forms = {}, {}
        quizzes = Quiz.objects.filter(created_by = self.request.user).order_by('-created_date')
        for quiz in quizzes:
            self.quizzes[quiz.id] = quiz
            db_ids.append(quiz.id)
        button = self.request.POST['submit']
        if button is None:
            self.raise_non_field_error("Invalid button")
            return
        if button not in ['Add', 'Back', 'Cancel', 'Check', 'Preview', 'Save', 'Start', 'Update', 'Upload']:
            self.raise_non_field_error("Invalid button")
            return
        quiz_pk = self.request.POST.get('button-id')
        if quiz_pk is not None:
            quiz_pk = int(quiz_pk)
        if button in ['Check', 'Preview', 'Start', 'Update', 'Upload'] and quiz_pk is None:
            self.raise_non_field_error("Invalid button")
            return
        if button in ['Preview'] and self.quizzes[quiz_pk].questions_in_quiz.count() == 0:
            self.raise_non_field_error("No questions to preview")
            return
        if button in ['Start'] and self.quizzes[quiz_pk].status.description != cs.READY:
            self.raise_non_field_error("Invalid status for hosting quiz")
            return
        if button in ['Start']:
            quiz = Quiz.objects.get_or_none(pk=quiz_pk)
            room = QuizRoom.objects.get_or_none(
                quiz=quiz, host=quiz.created_by, opened_date = dt.today().astimezone(), preview=False)
            status = RoomMemberStatus.objects.get_or_none(description=rms.JOINED)
            if room is not None:
                quizRoomHost = QuizRoomMember.objects.get_or_none(user=self.request.user, room=room, status=status)
                if quizRoomHost is not None:
                    self.raise_non_field_error("Quiz already being hosted")
                    return
        for form in self.forms:
            if len(form.field_custom_errors) > 0:
                self.field_custom_errors = form.field_custom_errors
                return
            if int(form.cleaned_data['id'].id) not in db_ids:
                self.raise_non_field_error("Invalid id")
                return
            self.received_forms[form.cleaned_data['id'].id] = form
            received_ids.append(form.cleaned_data['id'].id)
        if received_ids != db_ids:
            self.raise_non_field_error("Ids don't match error")
            return

        if len(received_ids) != len(set(received_ids)):
            self.raise_non_field_error("Duplicate ids")
            return

    def delete_media(self, id):
        fs = FileSystemStorage()
        fs.location += f"/{id}"
        if fs.exists(fs.location):
            files = fs.listdir(fs.location)
            for file in files[1]:
                fs.delete(name=file)
            fs.delete(name=fs.location)

    # def get_error_messages(self, formset, resp):
    #     if len(formset.field_custom_errors) == 0:
    #         for form in formset:
    #             if len(form.field_custom_errors) > 0:
    #                 resp['errors'] = form.field_custom_errors
    #                 break
    #     else:
    #         resp['errors'] = formset.field_custom_errors
    #     if not 'errors' in resp:
    #         resp['errors'] = {'success_message': 'Unknown error encountered'}

    @staticmethod
    def get_formset(request):
        formData = request.POST.get('form')
        if formData is None:
            return None
        decodedFormData = dict(urllib.parse.parse_qsl(formData))
        formset = ListQuizFormSet(decodedFormData, request=request)
        return formset

    def process_add(self, request, resp):
        user = User.objects.get(username=request.user.username)
        quiz = Quiz(
            created_by=user, title="New Quiz")
        quiz.save()
        request.session['message'] = {"message": "Updates Saved", "message_type": mt.SUCCESS}
        resp['redirect'] = reverse("synaptic:list_quizzes")

    def process_back(self, request, resp):
        request.session['quiz_pk'] = None
        request.session['question_action'] = None
        resp['redirect'] = reverse("synaptic:index")

    def process_button(self, request, resp):
        button = request.POST['submit']
        if button == "Add":
            self.process_add(request, resp)
        elif button == "Back":
            self.process_back(request, resp)
        elif button == "Check":
            request.session['quiz_pk'] = int(request.POST.get('button-id'))
            self.process_check(request, resp)
        elif button == "Preview":
            self.process_preview(request, resp)
        elif button == "Start":
            self.process_start(request, resp)
        elif button == "Update":
            self.process_update(request, resp)
        elif button == "Upload":
            self.process_upload(request, resp)

    @staticmethod
    def process_cancel(resp):
        resp['redirect'] = reverse("synaptic:list_quizzes")
        resp['success'] = True

    @staticmethod
    def process_check(request, resp):
        success_status = CheckStatus.objects.get_or_none(description=cs.READY)
        failed_status = CheckStatus.objects.get_or_none(description=cs.NOT_READY)
        quiz = Quiz.objects.get_or_none(pk=request.session['quiz_pk'])
        questions = Question.objects.filter(quiz=quiz).exclude(status=success_status).all()
        request.session['Check'] = "Check"
        failed_questions = False
        for question in questions:
            initial_values=question.get_fields()
            form = QuestionForm(data=initial_values, request=request)
            if form.is_valid():
                question.status = success_status
            else:
                question.status = failed_status
                failed_questions = True
            question.save()
        del request.session['Check']
        if failed_questions:
            quiz.status = failed_status
            request.session['message'] = {"message": "Some questions need attention", "message_type": mt.ERROR}
            resp['redirect'] = reverse("synaptic:list_questions")
            request.session['return_to_url'] = "list_quizzes"
        else:
            quiz.status = success_status
            request.session['message'] = {"message": "Check completed successfully", "message_type": mt.SUCCESS}
            resp['redirect'] = reverse("synaptic:list_quizzes")
        quiz.save()

    def process_preview(self, request, resp):
        room = generate_room(request.POST.get('button-id'), status=rs.QUESTION_PREVIEW, preview=True)
        request.session['preview_return_to_url'] = "list_quizzes"
        resp['redirect'] = reverse("synaptic:live_room", args=[room.room_number])

    def process_start(self, request, resp):
        #generate_room(quiz_id, status=rs.WAITING, preview=False, first_question_number=None, last_question_number=None)
        room = generate_room(request.POST['button-id'])
        request.session['return_to_url'] = "list_quizzes"
        resp['redirect'] = reverse("synaptic:live_room", args=[room.room_number])

    def process_update(self, request, resp):
        request.session['quiz_pk'] = int(request.POST.get('button-id'))
        request.session['question_action'] = ff.UPDATE
        resp['redirect'] = reverse("synaptic:list_questions")

    def process_upload(self, request, resp):
        request.session['quiz_pk'] = int(request.POST.get('button-id'))
        request.session['question_action'] = ff.UPDATE
        resp['redirect'] = reverse("synaptic:upload_quiz_spreadsheet")

    def save_changes(self, request, resp):
        for form in list(self.received_forms.values()):
            quiz = self.quizzes[form.cleaned_data['id'].id]
            if form.cleaned_data['deleted'] == True:
                quiz.delete()
                self.delete_media(form.cleaned_data['id'].id)
                resp['redirect'] = reverse("synaptic:list_quizzes")
                request.session['message'] = {"message": "Updates saved", "message_type": mt.SUCCESS}
                continue
            if quiz.title == form.cleaned_data['title']:
                continue
            quiz.title = form.cleaned_data['title']
            quiz.save()
            request.session['message'] = {"message": "Updates saved", "message_type": mt.SUCCESS}
            resp['redirect'] = reverse("synaptic:list_quizzes")

ListQuizFormSet = modelformset_factory(Quiz, formset=BaseListQuizFormSet, form=ListQuizForm, extra=0)

class QuestionForm(ModelForm):
    file_upload_media = forms.FileField(required=False, widget=forms.FileInput(attrs={
        "class": "file_upload",
        "multiple": False,
        "style": "display: none",
        'accept': "images/*",
        'onchange': "receiveFilesMedia()"
    }))
    class Meta:
        model = Question;
        fields = ['question', 'transition_type', 'media_url', 'time_limit', 'score_multiplier',
                  'answer_1', 'answer_2', 'answer_3', 'answer_4',
                  'correct_answer_1', 'correct_answer_2', 'correct_answer_3', 'correct_answer_4']
        widgets = {
            'question': Textarea(
                attrs={"rows": 1, "placeholder": "Question text",
                       "class": "border border-primary primary-font font-colour-pri text-lg form-control auto-resize"}),
            'media_url': Textarea(
                attrs={"rows": 1, "placeholder": "Media name/url",
                       "class": "border border-primary primary-font font-colour-pri text-md form-control auto-resize"}),
            'transition_type': Select(
                attrs={"class": "border border-primary primary-font font-colour-pri text-md form-control"}),
            'answer_1': Textarea(
                attrs={"rows": 1,
                       "class": "form-control rounded auto-resize-answer-buttons align-self-center answer-button-text answer-1 ms-0 me-0 ps-o pe-0",
                       "data-container_id": "#answer-1-container", "data-font_size": "text-xl", "data-shape": "rounded-pill"}),
            'answer_2': Textarea(
                attrs={"rows": 1,
                       "class": "form-control rounded auto-resize-answer-buttons align-self-center answer-button-text answer-2 ms-0 me-0 ps-o pe-0",
                       "data-container_id": "#answer-2-container", "data-font_size": "text-xl", "data-shape": "rounded-pill"}),
            'answer_3': Textarea(
                attrs={"rows": 1,
                       "class": "form-control rounded auto-resize-answer-buttons align-self-center answer-button-text answer-3  ms-0 me-0 ps-o pe-0",
                       "data-container_id": "#answer-3-container", "data-font_size": "text-xl", "data-shape": "rounded-pill"}),
            'answer_4': Textarea(
                attrs={"rows": 1,
                       "class": "form-control rounded auto-resize-answer-buttons align-self-center answer-button-text answer-4 ms-0 me-0 ps-o pe-0",
                       "data-container_id": "#answer-4-container", "data-font_size": "text-xl", "data-shape": "rounded-pill"}),
            'time_limit': TextInput(attrs={"class": "border border-primary form-control integer", "maxlength": 3}),
            'score_multiplier': TextInput(attrs={"class": "border border-primary form-control multiplier"}),
        }
    def __init__(self, *args, **kwargs):
        self.request = kwargs.pop('request', None)
        super().__init__(*args, **kwargs)
        self.fields['transition_type'].empty_label = "Transition"
        self.non_field_alert = {}
        self.field_custom_errors = {}
        self.upload_filename = None
        self.di = di()

    def clean(self):
        self.cleaned_data = super().clean()
        if not len(self.errors) == 0:
            self.validation_passed = False
            return

        button = self.request.POST.get('submit')
        question_action = self.request.session.get("question_action")
        if not self.has_changed() and button == "Back":
            return

        validate = False
        if button == "Check" or self.request.session.get('Check') == "Check":
            validate = True

        self.validation_passed = True

        if self.cleaned_data['question'] == "":
            add_custom_field_error(self, 'question', 'Enter question text', '#id_question_error')
            self.validation_passed = False

        upload_filenames = get_upload_media_filenames(self.request, 'file_upload')
        if len(upload_filenames) > 1:
            add_custom_field_error(self, 'media_url', 'Cannot recognise media url filename', '#id_media_url_error')
            self.validation_passed = False
        else:
            if len(upload_filenames) == 1:
                self.upload_filename = upload_filenames[0]

        media_url = self.cleaned_data.get('media_url')
        if media_url is not None and media_url != self.upload_filename:
            try:
                #r = requests.get(self.cleaned_data['media_url'], verify=False)
                if media_url.startswith('http'):
                    content_type = get_url_content_type(media_url)
                    if content_type not in ['image/jpg', 'image/jpeg', 'image/gif']:
                        add_custom_field_error(self, 'media_url', 'Content at url is not a valid image', '#id_media_url_error')
                        self.validation_passed = False
                else:
                    fs = FileSystemStorage()
                    quiz_pk = self.request.session.get('quiz_pk', None)
                    fs.location += f"/{quiz_pk}"
                    if not fs.exists(media_url):
                        add_custom_field_error(self, 'media_url', 'Cannot find image in library - add media', '#id_media_url_error')
                        self.validation_passed = False
            except:
                add_custom_field_error(self, 'media_url', 'Cannot open image at url given', '#id_media_url_error')
                self.validation_passed = False

        if self.cleaned_data['time_limit'] <= 0:
            add_custom_field_error(self, 'time_limit', 'Time limit must be greater than 0', '#id_time_limit_error')
            self.validation_passed = False

        if self.cleaned_data['score_multiplier'] > 4:
            add_custom_field_error(self, 'score_multiplier', 'Score multiplier must be 4 or less', '#id_score_multiplier_error')
            self.validation_passed = False
        elif self.cleaned_data['score_multiplier'] <= 0 :
            add_custom_field_error(self, 'score_multiplier', 'Score multiplier must be greater than 0', '#id_score_multiplier_error')
            self.validation_passed = False

        correct_answer_count = 0
        answer_error = False
        blank_answers = []
        last_non_blank_answer = 0
        for i in range(0, 4):
            answer = self.cleaned_data[f"answer_{i+1}"]
            correct_answer = self.cleaned_data[f"correct_answer_{i+1}"]
            if correct_answer == True:
                correct_answer_count += 1
            if answer == None or answer == "":
                blank_answers.append(i+1)
                if correct_answer == True:
                    if validate:
                        add_custom_field_error(self, f'answer_{i+1}', 'Blank answer marked as correct', f'#id_answer_{i+1}_error')
                    self.validation_passed = False
                    answer_error = True
            else:
                last_non_blank_answer = i+1

        if not answer_error:
            for answer in [answer for answer in blank_answers if answer < last_non_blank_answer]:
                if validate:
                    add_custom_field_error(self, f'answer_{answer}', 'Gap in answers entered', f'#id_answer_{answer}_error')
                self.validation_passed = False
                answer_error = True

        if not answer_error:
            if correct_answer_count == 0:
                for i in range(0, 4):
                    if validate and not isEmpty(self.cleaned_data[f"answer_{i+1}"]):
                        add_custom_field_error(self, f'answer_{i+1}', 'No correct answer marked', f'#id_answer_{i+1}_error')
                    self.validation_passed = False
                    answer_error = True
            elif last_non_blank_answer == 1:
                if validate:
                    add_custom_field_error(self, 'answer_2', 'Only 1 answer entered', '#id_answer_2_error')
                self.validation_passed = False
                answer_error = True

    def has_validation_changed(self, question):
        validation_changed = False
        if question is not None:
            if question.status.description == cs.READY and self.validation_passed == False:
                validation_changed = True
            if question.status.description == cs.NOT_READY and self.validation_passed == True:
                validation_changed = True
        return validation_changed

    @staticmethod
    def initialise_create(quiz, request):
        if request.session["question_number"] is None:
            request.session["question_number"] = get_new_question_number(request, quiz)
            form = QuestionForm()
        else:
            question = Question.objects.get_or_none(
                quiz=quiz, question_number = request.session["question_number"])
            # - to be deleted - doesn't appear to be doing anything - will only change if question is None
            # request.session["question_number"] = question.question_number
            form = QuestionForm(instance=question)
        icon_buttons = get_icon_buttons(
            add=True, add_button_text="Save & Add New Question", add_button_state="enabled",
            check=True, check_button_state="enabled"
        )
        return request, form, icon_buttons

    @staticmethod
    def initialise_update(quiz, request):
        question = Question.objects.get_or_none(
            quiz=quiz, question_number = request.session["question_number"])
        # - to be deleted - doesn't appear to be doing anything - will only change if question is None
        # request.session["question_number"] = question.question_number
        form = QuestionForm(instance=question)

        prev_button_state = ""
        if question.question_number == 1:
            prev_button_state = "disabled"
        last_question_number = Question.objects.filter(quiz=quiz).aggregate(Max('question_number'))
        next_button_state = ""
        if question.question_number == last_question_number['question_number__max']:
            next_button_state = "disabled"

        icon_buttons = get_icon_buttons(
            next=True, prev=True, next_button_state=next_button_state, prev_button_state=prev_button_state,
            check=True, check_button_state="enabled"
        )
        return request, form, icon_buttons

    def process_add_button(self, quiz, request, resp):
        last_question_number = get_last_question_number(quiz)
        request.session['question_number'] = 1
        if last_question_number is not None:
            request.session['question_number'] = last_question_number + 1
        resp['redirect'] = reverse("synaptic:question")
        resp['success'] = True

    def process_back_button(self, resp):
        resp['redirect'] = reverse("synaptic:list_questions")
        resp['success'] = True

    @staticmethod
    def process_cancel_button(resp):
        resp['redirect'] = reverse("synaptic:question")
        resp['success'] = True

    def process_next_button(self, quiz, request, resp):
        last_question_number = get_last_question_number(quiz)
        if last_question_number != None:
            if request.session['question_number'] < last_question_number:
                request.session['question_number'] += 1
        resp['redirect'] = reverse("synaptic:question")
        resp['success'] = True

    def process_prev_button(self, request, resp):
        if request.session['question_number'] > 1:
            request.session['question_number'] -= 1
        resp['redirect'] = reverse("synaptic:question")
        resp['success'] = True

    def process_save_button(self, request, resp):
        #resp['success-message'] = request.session['message']['message']
        #request.session['message'] = {"message": "", "message_type": mt.SUCCESS}
        resp['redirect'] = reverse("synaptic:question")
        resp['success'] = True

    def save_question(self, quiz, question, request, resp):
        # Bypass save if the form has not changed and is creating a new question
        # (This avoids getting stuck in the screen with an error and being unable to exit)
        # Will still perform save in update mode if form has not changed but the status is 'not checked'
        # so that the correct status will be set
        button = self.request.POST.get('submit')
        file_uploaded = False
        if self.request.FILES.get('file_upload') is not None:
            file_uploaded = True

        if not self.has_changed() and not file_uploaded and question is None:
            return

        # Don't create new question if back button is pressed when creating a new question and no changes
        # have been made
        question_action = self.request.session.get("question_action")
        if not self.has_changed() and not file_uploaded and button == "Back":
            return

        if not self.has_changed() and not file_uploaded and not self.has_validation_changed(question):
            return
        upload_media_files(request, quiz, 'file_upload')
        model_fields = self.cleaned_data
        model_fields.pop('file_upload_media')
        if model_fields['transition_type'] == None:
            transition = TransitionType.objects.get_or_none(function=at.HORIZONTAL_GROW)
            model_fields['transition_type'] = transition
        status = cs.NOT_READY
        if self.validation_passed:
            status = cs.READY
        check_status = CheckStatus.objects.get_or_none(description = status)
        model_fields['status'] = check_status
        model_fields['default_image_number'] = self.di.get_random_default_image_number()
        question, created = Question.objects.get_or_create(
            quiz=quiz, question_number=request.session["question_number"], defaults=model_fields)
        if not created:
            for attr, value in model_fields.items():
                setattr(question, attr, value)
            question.save()
        self.update_quiz_status(quiz, check_status)

        if self.has_changed():
            request.session['message'] = {"message": "Updates saved", "message_type": mt.SUCCESS}
        request.session["question_number"] = question.question_number

    def update_quiz_status(self, quiz, check_status):
        total_questions = Question.objects.filter(quiz=quiz).count()
        unpassed_questions = Question.objects.filter(quiz=quiz).exclude(status__description=cs.READY).count()
        update_quiz_status = False
        if total_questions > 0 and check_status.description == cs.READY and unpassed_questions == 0:
            update_quiz_status = True
        if check_status.description != cs.READY:
            update_quiz_status = True
        if update_quiz_status:
            quiz.status = check_status
            quiz.save()

    @staticmethod
    def validate_button(request, resp):
        button = request.POST.get('submit')
        if button not in ['Add', 'Back', 'Cancel', 'Check', 'Next', 'Prev', 'Save']:
            resp = add_custom_non_field_alert(resp, _("Invalid submission received"))
            return resp

        question_action = request.session.get("question_action")
        if button in ['Next', 'Prev'] and question_action != ff.UPDATE:
            resp = add_custom_non_field_alert(resp, _("Invalid submission received"))
            return resp

        question_action = request.session.get("question_action")
        if button in ['Add'] and question_action != ff.CREATE:
            resp = add_custom_non_field_alert(resp, _("Invalid submission received"))
            return resp

        return resp

class UploadFilesForm(forms.Form):
    file_upload_excel = forms.FileField(required=False, widget=forms.FileInput(attrs={
        "class": "file_upload",
        "style": "display: none",
        'accept': "application/vnd.ms-excel, application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        'onchange': "receiveFilesExcel()"
    }))
    file_upload_media = forms.FileField(required=False, widget=MultipleFileInput(attrs={
        "class": "file_upload",
        "style": "display: none",
        'accept': "image/*",
        'onchange': "receiveFilesMedia()"
    }))
    def __init__(self, *args, **kwargs):
        self.request = kwargs.pop('request', None)
        super().__init__(*args, **kwargs)
        self.non_field_alert = {}
        self.field_custom_errors = {}
        self.di = di()
    def add_custom_field_error(self, field, error, target_css_id):
        # used to add errors when the errors are going to be passed to and rendered by js
        # field must be a valid field for the form
        # error is the error message to be displayed
        # target_field is the css id of the field to which the error message will be appended
        add_custom_field_error(self, field, error, target_css_id)
        self.add_error(field, error)
    def clean(self):
        self.cleaned_data = super().clean()
        if not len(self.errors) == 0:
            return

        if 'file_upload_excel' not in self.request.FILES:
        #    self.add_custom_field_error('file_upload_excel', 'Question spreadsheet must be attached', 'xl_dropbox')
        #    self.add_error('file_upload_excel', _('Question spreadsheet must be attached'))
            return

        excel_file = self.request.FILES['file_upload_excel']
        valid_file_types = ["application/vnd.ms-excel",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]
        if excel_file.content_type not in valid_file_types:

            self.add_custom_field_error('file_upload_excel', 'Attached file is not a valid spreadsheet', '#id_file_upload_excel_error')
            return

        file_extension = os.path.splitext(excel_file.name)[-1]
        valid_file_extensions = ['.xls', '.xlsx']
        if file_extension not in valid_file_extensions:
            self.add_custom_field_error(
                'file_upload_excel', 'Attached file is not a valid spreadsheet', '#id_file_upload_excel_error')
            return

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.worksheets[0]
        except:
            self.add_custom_field_error(
                'file_upload_excel', 'Could not open attached spreadsheet', '#id_file_upload_excel_error')
            return
        valid_headers = True
        for i in range(xl.headers_col, xl.headers_col + len(xl.headers)):
            if not xl.headers[i-1].get("mandatory"):
                continue
            header = ws[xl.headers_row][i].value
            if header is None or header == "":
                valid_headers = False
                break
            if not header.lower().startswith(xl.headers[i-1].get("title")):
                valid_headers = False
                break
        if not valid_headers:
            self.add_custom_field_error(
                'file_upload_excel', 'Incorrect headers detected in attached spreadsheet', '#id_file_upload_excel_error'
            )
            return

    def get_media_url(self, row, spreadsheet_headers, reference_headers):
        media_url = row[xl.media_url].value
        if isEmpty(spreadsheet_headers[xl.media_url].value):
            return None, self.di.get_random_default_image_number()

        if spreadsheet_headers[xl.media_url].value.lower().startswith(reference_headers[xl.media_url - 1].get("title")):
            if isEmpty(media_url):
                return None, self.di.get_random_default_image_number()
            return row[xl.media_url].value, None

        return None, self.di.get_random_default_image_number()

    def get_score_multiplier(self, row, spreadsheet_headers, reference_headers):
        score_multiplier = row[xl.score_multiplier].value
        if isEmpty(spreadsheet_headers[xl.score_multiplier].value):
            return 1

        if spreadsheet_headers[xl.score_multiplier].value.lower().startswith(reference_headers[xl.score_multiplier - 1].get("title")):
            return set_score_multiplier(row[xl.score_multiplier].value)

        return 1
    
    def excel_to_db(self, request):
        excel_file = request.FILES['file_upload_excel']
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.worksheets[0]
        xl_data = [row for row in ws.rows]
        quiz_pk=request.session.get('quiz_pk', None)
        quiz = Quiz.objects.get_or_none(pk=quiz_pk)
        question_number = 1
        if quiz is not None:
            if Question.objects.filter(quiz=quiz).count() > 0:
                question = Question.objects.filter(quiz=quiz).order_by('-question_number')[0]
                question_number = question.question_number + 1
        questions = []
        reference_headers = xl.headers
        transition_type = TransitionType.objects.get_or_none(function=at.HORIZONTAL_GROW)
        for i, row in enumerate(xl_data):
            if i < xl.headers_row - 1:
                continue
            if i == xl.headers_row - 1:
                spreadsheet_headers = row
            if i > xl.headers_row - 1 and row[xl.question].value is not None:
                time_limit = set_time_limit(row[xl.time_limit].value)
                media_url, default_image_number = self.get_media_url(row, spreadsheet_headers, reference_headers)
                score_multiplier = self.get_score_multiplier(row, spreadsheet_headers, reference_headers)
                correct_answers = str(row[xl.correct_answers].value).replace(" ","").split(",")
                correct_answers_dict = {answer: True for answer in correct_answers}

                questions.append(Question(
                    quiz=quiz, question=xl_data[i][xl.question].value, question_number=question_number,
                    time_limit=time_limit, score_multiplier=score_multiplier,
                    media_url=media_url, default_image_number=default_image_number,
                    transition_type=transition_type,
                    answer_1 = row[xl.answer1].value, correct_answer_1=correct_answers_dict.get("1", False),
                    answer_2 = row[xl.answer2].value, correct_answer_2=correct_answers_dict.get("2", False),
                    answer_3 = row[xl.answer3].value, correct_answer_3=correct_answers_dict.get("3", False),
                    answer_4 = row[xl.answer4].value, correct_answer_4=correct_answers_dict.get("4", False),
                ))

                question_number += 1
        objs = Question.objects.bulk_create(questions)

        resp={}
        BaseListQuizFormSet.process_check(request, resp)

    @staticmethod
    def process_cancel(resp):
        resp['redirect'] = reverse("synaptic:upload_quiz_spreadsheet")
        resp['success'] = True

    def save_changes(self, request, resp):
        quiz_pk = request.session.get('quiz_pk')
        fs = FileSystemStorage()
        fs.location += f"/{quiz_pk}"
        if request.FILES.get('file_upload_media') != None:
            for file in request.FILES.getlist('file_upload_media'):
                filename = fs.save(file.name, file)
            request.session['message'] = {"message": "Updates saved", "message_type": mt.SUCCESS}
        if request.FILES.get('file_upload_excel') != None:
            self.excel_to_db(request)
            request.session['message'] = {"message": "Updates saved", "message_type": mt.SUCCESS}
        resp['redirect'] = reverse("synaptic:upload_quiz_spreadsheet")

